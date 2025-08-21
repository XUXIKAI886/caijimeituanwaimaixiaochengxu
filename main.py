# -*- coding: utf-8 -*-
"""
呈尚策划 美团外卖数据工具
简单可靠的桌面应用程序

⚠️ 重要声明：仅供公司内部测试使用，禁止商用
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import json
import re
import requests
import asyncio
import aiohttp
import aiofiles
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
from pathlib import Path
from datetime import datetime
import threading
import logging
from urllib.parse import urlparse
from PIL import Image
import io
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import time

class FileChangeHandler(FileSystemEventHandler):
    """文件变化处理器"""
    
    def __init__(self, collector, file1_path, file2_path):
        super().__init__()
        self.collector = collector
        self.file1_path = os.path.abspath(file1_path)
        self.file2_path = os.path.abspath(file2_path)
        
    def on_modified(self, event):
        """文件修改事件处理"""
        if event.is_directory:
            return
            
        file_path = os.path.abspath(event.src_path)
        current_time = time.time()
        
        # 检查是否为第一份文档
        if file_path == self.file1_path:
            if current_time - self.collector.last_file1_mtime > 3:
                self.collector.last_file1_mtime = current_time
                self.collector.log_message(f"检测到第一份文档更新: {file_path}")
                
                # 在主线程中执行处理
                self.collector.root.after(100, lambda: self.collector.handle_file_update("file1"))
        
        # 检查是否为第二份文档
        elif file_path == self.file2_path:
            if current_time - self.collector.last_file2_mtime > 3:
                self.collector.last_file2_mtime = current_time
                self.collector.log_message(f"检测到第二份文档更新: {file_path}")
                
                # 在主线程中执行处理
                self.collector.root.after(100, lambda: self.collector.handle_file_update("file2"))

class MeituanDataCollector:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("呈尚策划 美团外卖数据工具")
        self.root.geometry("800x600")
        self.root.resizable(True, True)
        
        # 数据存储
        self.products = []
        self.categories = {}
        
        # 文件路径
        self.output_dir = Path("输出文件")
        self.output_dir.mkdir(exist_ok=True)
        
        # 设置日志
        self.setup_logging()
        
        # 文件监控相关
        self.observer = None
        self.is_monitoring = False
        self.last_file1_mtime = 0
        self.last_file2_mtime = 0
        self.auto_monitoring = True  # 启用自动监控
        
        # 创建界面
        self.create_widgets()
        
    def setup_logging(self):
        """设置日志"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('采集日志.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
    def create_widgets(self):
        """创建界面组件"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # 免责声明
        disclaimer_frame = ttk.Frame(main_frame)
        disclaimer_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        disclaimer_frame.configure(relief='solid', borderwidth=1)
        
        disclaimer_label = ttk.Label(disclaimer_frame, 
                                   text="⚠️ 重要声明：仅供公司内部测试使用，禁止商用", 
                                   font=("微软雅黑", 10, "bold"), 
                                   foreground="red")
        disclaimer_label.pack(pady=5)
        
        # 数据文件信息显示
        info_frame = ttk.LabelFrame(main_frame, text="数据源信息", padding="10")
        info_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 10))
        info_frame.columnconfigure(1, weight=1)
        
        # 显示默认文件路径
        ttk.Label(info_frame, text="店铺分类文件:").grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Label(info_frame, text="D:\\ailun\\xiaochengxumeituan.txt", font=("微软雅黑", 9)).grid(row=0, column=1, sticky=tk.W, padx=(10, 0))
        
        ttk.Label(info_frame, text="产品数据文件:").grid(row=1, column=0, sticky=tk.W, pady=2)
        ttk.Label(info_frame, text="D:\\ailun\\xiaochengxumeituan01.txt", font=("微软雅黑", 9)).grid(row=1, column=1, sticky=tk.W, padx=(10, 0))
        
        # 设置默认文件路径
        self.file1_path = "D:\\ailun\\xiaochengxumeituan.txt"
        self.file2_path = "D:\\ailun\\xiaochengxumeituan01.txt"
        
        # 控制按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, columnspan=3, pady=(20, 10))
        
        ttk.Button(button_frame, text="开始采集数据", command=self.start_collection, style="Accent.TButton").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="手动启动监控", command=self.start_monitoring).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="停止监控", command=self.stop_monitoring).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="打开输出文件夹", command=self.open_output_folder).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="清空日志", command=self.clear_log).pack(side=tk.LEFT)
        
        # 进度条
        self.progress_var = tk.StringVar(value="准备就绪")
        ttk.Label(main_frame, textvariable=self.progress_var, font=("微软雅黑", 10)).grid(row=3, column=0, columnspan=3, pady=(10, 5))
        
        self.progress_bar = ttk.Progressbar(main_frame, mode='determinate')
        self.progress_bar.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 统计信息
        stats_frame = ttk.LabelFrame(main_frame, text="统计信息", padding="10")
        stats_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        stats_frame.columnconfigure(1, weight=1)
        stats_frame.columnconfigure(3, weight=1)
        
        ttk.Label(stats_frame, text="产品数量:").grid(row=0, column=0, sticky=tk.W)
        self.product_count_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.product_count_var, font=("微软雅黑", 10, "bold")).grid(row=0, column=1, sticky=tk.W, padx=(5, 20))
        
        ttk.Label(stats_frame, text="下载图片:").grid(row=0, column=2, sticky=tk.W)
        self.image_count_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.image_count_var, font=("微软雅黑", 10, "bold")).grid(row=0, column=3, sticky=tk.W, padx=(5, 0))
        
        # 日志区域
        log_frame = ttk.LabelFrame(main_frame, text="处理日志", padding="5")
        log_frame.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 0))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(6, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, wrap=tk.WORD)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 绑定关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
    def log_message(self, message):
        """添加日志消息"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        
        # 在主线程中更新UI
        self.root.after(0, lambda: self._update_log_text(log_entry))
        
        # 同时记录到日志文件
        self.logger.info(message)
        
    def _update_log_text(self, text):
        """更新日志文本（在主线程中调用）"""
        self.log_text.insert(tk.END, text)
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def clear_log(self):
        """清空日志"""
        self.log_text.delete(1.0, tk.END)
        
    def update_progress(self, value, text=""):
        """更新进度"""
        self.root.after(0, lambda: self._update_progress_ui(value, text))
        
    def _update_progress_ui(self, value, text):
        """更新进度UI（在主线程中调用）"""
        self.progress_bar['value'] = value
        if text:
            self.progress_var.set(text)
        self.root.update_idletasks()
        
    def update_stats(self, products=None, images=None):
        """更新统计信息"""
        if products is not None:
            self.root.after(0, lambda: self.product_count_var.set(str(products)))
        if images is not None:
            self.root.after(0, lambda: self.image_count_var.set(str(images)))
            
    def start_collection(self):
        """开始数据采集"""
        # 检查文件是否存在
        file1 = self.file1_path
        file2 = self.file2_path
        
        if not os.path.exists(file1):
            messagebox.showerror("错误", f"数据文件不存在: {file1}")
            return
            
        if not os.path.exists(file2):
            messagebox.showerror("错误", f"数据文件不存在: {file2}")
            return
            
        # 在后台线程中执行采集
        thread = threading.Thread(target=self.collection_worker, args=(file1, file2))
        thread.daemon = True
        thread.start()
        
    def collection_worker(self, file1, file2):
        """数据采集工作线程"""
        try:
            self.log_message("开始数据采集...")
            self.update_progress(0, "正在解析数据...")
            
            # 步骤1: 完整解析第一份文档（分类+产品）
            self.log_message(f"正在完整解析第一份文档: {file1}")
            categories, products_file1 = self.parse_complete_data(file1)
            self.log_message(f"第一份文档：解析到 {len(categories)} 个分类，{len(products_file1)} 个产品")
            self.update_progress(25, "第一份文档解析完成")
            
            # 步骤2: 解析第二份文档的更新产品
            self.log_message(f"正在解析第二份文档的更新数据: {file2}")
            products_file2 = self.parse_products(file2, categories)
            self.log_message(f"第二份文档：解析到 {len(products_file2)} 个更新产品")
            self.update_progress(40, "第二份文档解析完成")
            
            # 步骤3: 合并所有产品数据
            all_products = products_file1 + products_file2
            self.log_message(f"合并数据：总共 {len(all_products)} 个产品")
            self.update_stats(products=len(all_products))
            
            # 步骤4: 下载图片
            if all_products:
                self.log_message("开始下载产品图片...")
                self.update_progress(50, "正在下载图片...")
                downloaded_count = asyncio.run(self.download_images(all_products))
                self.log_message(f"图片下载完成，成功下载 {downloaded_count} 张")
                self.update_progress(80, "图片下载完成")
                self.update_stats(images=downloaded_count)
                
                # 步骤5: 生成Excel报告
                self.log_message("正在生成Excel报告...")
                self.update_progress(90, "生成Excel报告")
                excel_path = self.generate_excel(all_products)
                self.log_message(f"Excel报告生成完成: {excel_path}")
                
            self.update_progress(100, "数据采集完成！")
            self.log_message("数据采集任务全部完成！")
            
            # 显示完成对话框
            self.root.after(0, lambda: messagebox.showinfo("完成", 
                f"数据采集完成！\n产品数量: {len(all_products)}\n输出目录: {self.output_dir}"))
                
        except Exception as e:
            error_msg = f"采集过程中出错: {str(e)}"
            self.log_message(error_msg)
            self.logger.error(error_msg, exc_info=True)
            self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
            self.update_progress(0, "采集失败")
            
    def parse_categories(self, file_path):
        """解析分类文件"""
        categories = {}
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
            data = json.loads(content)
            if 'data' in data and 'food_spu_tags' in data['data']:
                for category in data['data']['food_spu_tags']:
                    if 'name' in category and 'tag' in category:
                        categories[category['tag']] = category['name']
                        
        except Exception as e:
            self.log_message(f"解析分类文件错误: {e}")
            raise
            
        return categories
    
    def parse_complete_data(self, file_path):
        """完整解析文件数据（分类信息+产品信息）"""
        categories = {}
        products = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
            data = json.loads(content)
            if 'data' in data and 'food_spu_tags' in data['data']:
                # 第一轮：提取分类映射
                for category in data['data']['food_spu_tags']:
                    if 'name' in category and 'tag' in category:
                        categories[category['tag']] = category['name']
                
                # 第二轮：提取产品信息
                for category_group in data['data']['food_spu_tags']:
                    tag = category_group.get('tag', '')
                    category_name = categories.get(tag, '未知分类')
                    
                    # 检查不同的产品字段名
                    products_field = None
                    if 'spus' in category_group and category_group['spus']:
                        products_field = 'spus'
                    elif 'dynamic_spus' in category_group and category_group['dynamic_spus']:
                        products_field = 'dynamic_spus'
                    
                    if products_field and category_group[products_field]:
                        for product in category_group[products_field]:
                            product_info = {
                                'category_name': category_name,
                                'product_name': product.get('name', ''),
                                'product_price': product.get('min_price', 0),
                                'image_url': product.get('picture', ''),
                                'image_path': '',
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            products.append(product_info)
                            
        except Exception as e:
            self.log_message(f"完整解析文件错误: {e}")
            raise
            
        return categories, products
        
    def parse_products(self, file_path, categories):
        """解析产品文件"""
        products = []
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
            data = json.loads(content)
            if 'data' in data and 'food_spu_tags' in data['data']:
                for category in data['data']['food_spu_tags']:
                    # 获取分类名称
                    tag = category.get('tag', '')
                    category_name = categories.get(tag, '未知分类')
                    
                    # 检查不同的产品字段名
                    products_field = None
                    if 'spus' in category:
                        products_field = 'spus'
                    elif 'dynamic_spus' in category:
                        products_field = 'dynamic_spus'
                    
                    if products_field and category[products_field]:
                        for product in category[products_field]:
                            product_info = {
                                'category_name': category_name,
                                'product_name': product.get('name', ''),
                                'product_price': product.get('min_price', 0),  # 直接使用元为单位
                                'image_url': product.get('picture', ''),
                                'image_path': '',
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            products.append(product_info)
                            
        except Exception as e:
            self.log_message(f"解析产品文件错误: {e}")
            raise
            
        return products
        
    async def download_images(self, products):
        """下载产品图片"""
        images_dir = self.output_dir / "产品图片"
        images_dir.mkdir(exist_ok=True)
        
        downloaded_count = 0
        total_products = len(products)
        
        async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=30)) as session:
            for i, product in enumerate(products):
                try:
                    if not product['image_url']:
                        continue
                        
                    # 安全的文件名
                    safe_name = re.sub(r'[<>:"/\\|?*]', '_', product['product_name'])
                    image_path = images_dir / f"{safe_name}.jpg"
                    
                    # 如果文件已存在，跳过
                    if image_path.exists():
                        product['image_path'] = str(image_path)
                        downloaded_count += 1
                        continue
                        
                    # 下载图片
                    async with session.get(product['image_url']) as response:
                        if response.status == 200:
                            content = await response.read()
                            
                            # 转换为JPG格式
                            try:
                                image = Image.open(io.BytesIO(content))
                                if image.mode in ('RGBA', 'LA', 'P'):
                                    image = image.convert('RGB')
                                image.save(image_path, 'JPEG', quality=85)
                                
                                product['image_path'] = str(image_path)
                                downloaded_count += 1
                                
                            except Exception as img_error:
                                self.log_message(f"图片转换失败 {safe_name}: {img_error}")
                                
                except Exception as e:
                    self.log_message(f"下载图片失败 {product['product_name']}: {e}")
                    
                # 更新进度
                progress = 50 + int((i + 1) / total_products * 30)
                self.update_progress(progress, f"下载图片中... ({i+1}/{total_products})")
                
        return downloaded_count
        
    def generate_excel(self, products):
        """生成Excel报告"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_path = self.output_dir / f"美团外卖数据报告_{timestamp}.xlsx"
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "产品数据"
        
        # 设置表头
        headers = ['分类名称', '产品名称', '产品价格(元)', '图片URL', '本地图片路径', '采集时间']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        # 填充数据
        for row, product in enumerate(products, 2):
            sheet.cell(row=row, column=1, value=product['category_name'])
            sheet.cell(row=row, column=2, value=product['product_name'])
            sheet.cell(row=row, column=3, value=product['product_price'])
            sheet.cell(row=row, column=4, value=product['image_url'])
            sheet.cell(row=row, column=5, value=product['image_path'])
            sheet.cell(row=row, column=6, value=product['timestamp'])
            
        # 自动调整列宽
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
            
        # 保存文件
        workbook.save(excel_path)
        return excel_path.name
        
    def open_output_folder(self):
        """打开输出文件夹"""
        try:
            os.startfile(self.output_dir)
        except Exception as e:
            messagebox.showerror("错误", f"无法打开文件夹: {e}")
    
    def auto_start_monitoring(self):
        """自动启动文件监控"""
        file1 = self.file1_path
        file2 = self.file2_path
        
        if not file1 or not file2:
            self.log_message("请设置数据文件路径后，程序将自动启动监控")
            self.update_progress(0, "等待文件路径设置...")
            # 每5秒检查一次文件路径是否已设置
            self.root.after(5000, self.auto_start_monitoring)
            return
            
        if not os.path.exists(file1):
            self.log_message(f"等待第一份文档创建: {file1}")
            self.update_progress(0, "等待第一份文档...")
            self.root.after(5000, self.auto_start_monitoring)
            return
            
        if not os.path.exists(file2):
            self.log_message(f"等待第二份文档创建: {file2}")
            self.update_progress(0, "等待第二份文档...")
            self.root.after(5000, self.auto_start_monitoring)
            return
        
        # 如果监控已启动，则跳过
        if self.is_monitoring:
            return
            
        # 启动双文件监控
        self.start_dual_monitoring(file1, file2)
    
    def start_dual_monitoring(self, file1, file2):
        """启动双文件监控"""
        try:
            # 初始化文件修改时间
            self.last_file1_mtime = os.path.getmtime(file1)
            self.last_file2_mtime = os.path.getmtime(file2)
            
            # 创建观察者和处理器
            self.observer = Observer()
            handler = FileChangeHandler(self, file1, file2)
            
            # 监控两个文件所在的目录
            dir1 = os.path.dirname(os.path.abspath(file1))
            dir2 = os.path.dirname(os.path.abspath(file2))
            
            self.observer.schedule(handler, dir1, recursive=False)
            if dir1 != dir2:  # 如果是不同目录，也要监控第二个目录
                self.observer.schedule(handler, dir2, recursive=False)
            
            # 启动观察者
            self.observer.start()
            self.is_monitoring = True
            
            self.log_message("自动文件监控已启动")
            self.log_message(f"监控第一份文档: {file1}")
            self.log_message(f"监控第二份文档: {file2}")
            self.update_progress(0, "自动监控已启动")
            
        except Exception as e:
            error_msg = f"自动启动监控失败: {str(e)}"
            self.log_message(error_msg)
            self.update_progress(0, "监控启动失败")
    
    def start_monitoring(self):
        """手动启动监控（保留原功能）"""
        file2 = self.file2_path
        
        if not file2:
            messagebox.showerror("错误", "请选择第二份数据文件")
            return
            
        if not os.path.exists(file2):
            messagebox.showerror("错误", f"文件不存在: {file2}")
            return
        
        if self.is_monitoring:
            messagebox.showwarning("提示", "监控已在运行中")
            return
        
        try:
            # 初始化文件修改时间
            self.last_file2_mtime = os.path.getmtime(file2)
            
            # 创建观察者和处理器
            self.observer = Observer()
            handler = FileChangeHandler(self, file2)
            
            # 监控文件所在目录
            dir_path = os.path.dirname(os.path.abspath(file2))
            self.observer.schedule(handler, dir_path, recursive=False)
            
            # 启动观察者
            self.observer.start()
            self.is_monitoring = True
            
            self.log_message(f"开始监控第二份文档: {file2}")
            self.update_progress(0, "监控已启动")
            messagebox.showinfo("成功", "文件监控已启动！\n程序将自动检测第二份文档的更新")
            
        except Exception as e:
            error_msg = f"启动监控失败: {str(e)}"
            self.log_message(error_msg)
            messagebox.showerror("错误", error_msg)
    
    def stop_monitoring(self):
        """停止文件监控"""
        if not self.is_monitoring:
            self.log_message("监控未在运行")
            return
        
        try:
            if self.observer:
                self.observer.stop()
                self.observer.join(timeout=5)
                self.observer = None
            
            self.is_monitoring = False
            self.log_message("文件监控已停止")
            self.update_progress(0, "监控已停止")
            
        except Exception as e:
            error_msg = f"停止监控失败: {str(e)}"
            self.log_message(error_msg)
            self.logger.error(error_msg, exc_info=True)
    
    def handle_file_update(self, file_type):
        """统一处理文件更新"""
        try:
            file1 = self.file1_path
            file2 = self.file2_path
            
            if not file1 or not file2:
                self.log_message("文件路径未设置，跳过自动处理")
                return
            
            if not os.path.exists(file1) or not os.path.exists(file2):
                self.log_message("文件不存在，跳过自动处理") 
                return
            
            if file_type == "file1":
                self.log_message("检测到第一份文档更新，开始完整重新采集...")
                self.update_progress(0, "重新采集数据中...")
                # 在后台线程中执行完整采集
                thread = threading.Thread(target=self.full_collection_worker, args=(file1, file2))
            else:  # file2
                self.log_message("检测到第二份文档更新，开始增量采集...")
                self.update_progress(0, "增量采集中...")
                # 在后台线程中执行增量采集
                thread = threading.Thread(target=self.incremental_worker, args=(file1, file2))
            
            thread.daemon = True
            thread.start()
            
        except Exception as e:
            error_msg = f"处理文档更新时出错: {str(e)}"
            self.log_message(error_msg)
    
    def full_collection_worker(self, file1, file2):
        """完整采集工作线程（第一份文档更新时）"""
        try:
            self.log_message("开始完整数据重新采集...")
            
            # 步骤1: 完整解析第一份文档（分类+产品）
            categories, products_file1 = self.parse_complete_data(file1)
            self.log_message(f"第一份文档：解析到 {len(categories)} 个分类，{len(products_file1)} 个产品")
            self.update_progress(30, "第一份文档解析完成")
            
            # 步骤2: 解析第二份文档的更新产品
            products_file2 = self.parse_products(file2, categories)
            self.log_message(f"第二份文档：解析到 {len(products_file2)} 个产品")
            self.update_progress(50, "第二份文档解析完成")
            
            # 步骤3: 合并所有产品数据
            all_products = products_file1 + products_file2
            self.log_message(f"合并数据：总共 {len(all_products)} 个产品")
            self.update_stats(products=len(all_products))
            
            # 步骤4: 下载图片
            if all_products:
                self.log_message("开始下载产品图片...")
                self.update_progress(60, "正在下载图片...")
                downloaded_count = asyncio.run(self.download_images(all_products))
                self.log_message(f"图片下载完成，成功下载 {downloaded_count} 张")
                self.update_progress(85, "图片下载完成")
                self.update_stats(images=downloaded_count)
                
                # 步骤5: 生成Excel报告
                self.log_message("正在生成Excel报告...")
                self.update_progress(95, "生成Excel报告")
                excel_path = self.generate_excel(all_products)
                self.log_message(f"Excel报告生成完成: {excel_path}")
                
            self.update_progress(100, "完整采集完成")
            self.log_message("完整数据采集任务完成！")
            
        except Exception as e:
            error_msg = f"完整采集失败: {str(e)}"
            self.log_message(error_msg)
            self.logger.error(error_msg, exc_info=True)
            self.update_progress(0, "采集失败")
    
    def incremental_worker(self, file1, file2):
        """增量采集工作线程（第二份文档更新时）"""
        try:
            # 解析第一份文档获取分类信息
            categories, existing_products = self.parse_complete_data(file1)
            self.log_message(f"获取到 {len(categories)} 个分类信息")
            
            # 解析第二份文档的新数据
            new_products = self.parse_products(file2, categories)
            self.log_message(f"检测到 {len(new_products)} 个更新产品")
            
            if new_products:
                # 计算总产品数量（现有 + 新增）
                total_products = len(existing_products) + len(new_products)
                
                # 下载新产品的图片
                self.log_message("开始下载新产品图片...")
                self.update_progress(50, "下载新图片中...")
                downloaded_count = asyncio.run(self.download_images(new_products))
                self.log_message(f"新图片下载完成，成功下载 {downloaded_count} 张")
                self.update_progress(80, "新图片下载完成")
                
                # 追加到Excel（如果存在现有Excel文件）
                self.append_to_excel(new_products)
                
                # 更新统计信息：显示总的产品和图片数量
                self.update_stats(products=total_products)
                
                # 计算总的图片数量（需要检查实际已下载的图片）
                images_dir = self.output_dir / "产品图片"
                if images_dir.exists():
                    total_images = len(list(images_dir.glob("*.jpg")))
                    self.update_stats(images=total_images)
                
                self.update_progress(100, "增量更新完成")
                self.log_message(f"增量数据采集完成！新增产品: {len(new_products)} 个")
                
            else:
                self.log_message("未发现新的产品数据")
                self.update_progress(0, "无新数据")
                
        except Exception as e:
            error_msg = f"增量采集失败: {str(e)}"
            self.log_message(error_msg)
            self.logger.error(error_msg, exc_info=True)
            self.update_progress(0, "增量采集失败")
    
    def append_to_excel(self, new_products):
        """将新产品追加到Excel文件"""
        try:
            # 查找最新的Excel文件
            excel_files = list(self.output_dir.glob("美团外卖数据报告_*.xlsx"))
            if not excel_files:
                # 如果没有现有文件，创建新文件
                excel_path = self.generate_excel(new_products)
                self.log_message(f"创建新Excel文件: {excel_path}")
                return
            
            # 使用最新的Excel文件
            latest_excel = max(excel_files, key=lambda x: x.stat().st_mtime)
            self.log_message(f"追加数据到现有Excel文件: {latest_excel.name}")
            
            # 读取现有Excel文件
            workbook = openpyxl.load_workbook(latest_excel)
            sheet = workbook.active
            
            # 找到下一个可用行
            next_row = sheet.max_row + 1
            
            # 追加新产品数据
            for product in new_products:
                sheet.cell(row=next_row, column=1, value=product['category_name'])
                sheet.cell(row=next_row, column=2, value=product['product_name'])
                sheet.cell(row=next_row, column=3, value=product['product_price'])
                sheet.cell(row=next_row, column=4, value=product['image_url'])
                sheet.cell(row=next_row, column=5, value=product['image_path'])
                sheet.cell(row=next_row, column=6, value=product['timestamp'])
                next_row += 1
            
            # 保存文件
            workbook.save(latest_excel)
            self.log_message(f"成功追加 {len(new_products)} 行数据到Excel")
            
        except Exception as e:
            self.log_message(f"追加Excel数据失败: {e}")
            # 如果追加失败，创建新文件
            excel_path = self.generate_excel(new_products)
            self.log_message(f"创建新Excel文件: {excel_path}")
            
    def on_closing(self):
        """程序关闭事件"""
        # 停止文件监控
        if self.is_monitoring:
            self.stop_monitoring()
        
        self.root.quit()
        self.root.destroy()
        
    def run(self):
        """运行程序"""
        self.log_message("呈尚策划 美团外卖数据工具已启动")
        self.log_message("程序将自动监控数据文件变化")
        
        # 自动启动监控
        self.root.after(1000, self.auto_start_monitoring)  # 延迟1秒启动监控
        
        self.root.mainloop()

if __name__ == "__main__":
    app = MeituanDataCollector()
    app.run()